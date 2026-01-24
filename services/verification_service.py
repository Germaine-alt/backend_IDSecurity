from config.database import db
from models.lieu import Lieu
from models.verification import Verification
from flask_jwt_extended import get_jwt_identity
from datetime import datetime, timedelta
from sqlalchemy import func, extract
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime, timedelta
from sqlalchemy.orm import joinedload



class VerificationService:

    @staticmethod
    def save_verification(
        utilisateur_id: int,
        lieu_id: int,
        resultat_donnee: str,
        document_id: int = None,
        ocr_result_id: int = None,
        resultat_photo: str = "NON_VERIFIE",
        url_image_echec: str = None
    ) -> Verification:
        verification = Verification(
            utilisateur_id=utilisateur_id,
            lieu_id=lieu_id,
            document_id=document_id,
            ocr_result_id=ocr_result_id,
            resultat_donnee=resultat_donnee,
            resultat_photo=resultat_photo,
            url_image_echec=url_image_echec
        )
        db.session.add(verification)
        try:
            db.session.commit()
            db.session.refresh(verification)
            print(f"✅ Verification enregistrée avec ID: {verification.id}, document_id: {document_id}")
        except Exception as e:
            db.session.rollback()
            print(f"❌ Erreur lors de l'enregistrement de la vérification: {e}")
            raise
        return verification


    @staticmethod
    def get_verification_by_id(id):
        return Verification.query.get(id)

    # @staticmethod
    # def get_all_verifications():
    #     return Verification.query.all()


    @staticmethod
    def get_all_verifications():
        return Verification.query.options(
            joinedload(Verification.utilisateur),
            joinedload(Verification.lieu),
            joinedload(Verification.document),
            joinedload(Verification.ocr_results)
        ).order_by(
            Verification.date_verification.desc()
        ).all()

    @staticmethod
    def get_user_verifications():          
        current_user_id = get_jwt_identity()          
        verifications = Verification.query.filter_by(            
            utilisateur_id=current_user_id        
            ).order_by(Verification.date_verification.desc()).all()          
        return verifications    



    @staticmethod
    def get_statistiques_verifications(periode=None):
        """Récupère les statistiques avec filtrage par période"""
        query = Verification.query
        
        if periode:
            now = datetime.now()
            if periode == "today":
                # Aujourd'hui (minuit à maintenant)
                start_of_day = datetime.combine(now.date(), datetime.min.time())
                query = query.filter(Verification.date_verification >= start_of_day)
            elif periode == "yesterday":
                # Hier (toute la journée)
                yesterday = now - timedelta(days=1)
                start_of_day = datetime.combine(yesterday.date(), datetime.min.time())
                end_of_day = datetime.combine(yesterday.date(), datetime.max.time())
                query = query.filter(
                    Verification.date_verification >= start_of_day,
                    Verification.date_verification <= end_of_day
                )
            elif periode == "week":
                # 7 derniers jours
                week_ago = now - timedelta(days=7)
                query = query.filter(Verification.date_verification >= week_ago)
            elif periode == "month":
                # 30 derniers jours
                month_ago = now - timedelta(days=30)
                query = query.filter(Verification.date_verification >= month_ago)
            
        total = query.count()
        
        reussies = query.filter(
            Verification.resultat_donnee == "OK"
        ).count()

        echouees = query.filter(
            Verification.resultat_donnee == "ECHEC"
        ).count()

        externes = query.filter(
            Verification.resultat_donnee == "EXTERNE"
        ).count()

        return {
            "total": total,
            "reussies": reussies,
            "echouees": echouees,
            "externes": externes,
            "periode": periode or "all"
        }

    @staticmethod
    def get_stats_verifications_par_lieu(periode=None):
        """Récupère les statistiques par lieu avec filtrage par période"""
        query = Verification.query
        
        if periode:
            now = datetime.now()
            if periode == "today":
                start_of_day = datetime.combine(now.date(), datetime.min.time())
                query = query.filter(Verification.date_verification >= start_of_day)
            elif periode == "yesterday":
                yesterday = now - timedelta(days=1)
                start_of_day = datetime.combine(yesterday.date(), datetime.min.time())
                end_of_day = datetime.combine(yesterday.date(), datetime.max.time())
                query = query.filter(
                    Verification.date_verification >= start_of_day,
                    Verification.date_verification <= end_of_day
                )
            elif periode == "week":
                week_ago = now - timedelta(days=7)
                query = query.filter(Verification.date_verification >= week_ago)
            elif periode == "month":
                month_ago = now - timedelta(days=30)
                query = query.filter(Verification.date_verification >= month_ago)
            
        
        subquery = query.filter(
            Verification.lieu_id.isnot(None)
        ).with_entities(
            Verification.lieu_id,
            func.max(Verification.date_verification).label('derniere_verif')
        ).group_by(
            Verification.lieu_id
        ).subquery()

        results = db.session.query(
            Verification.lieu_id,
            func.count(Verification.id)
        ).join(
            subquery,
            Verification.lieu_id == subquery.c.lieu_id
        ).group_by(
            Verification.lieu_id,
            subquery.c.derniere_verif
        ).order_by(
            subquery.c.derniere_verif.desc()  
        ).limit(10).all()

        data = []
        for lieu_id, total in results:
            lieu = Lieu.query.get(lieu_id)
            if lieu:
                data.append({
                    "lieu": lieu.nom,
                    "total": total
                })

        return data

    @staticmethod
    def get_dernieres_verifications(periode=None, limit=4):
        """Récupère les dernières vérifications avec filtrage par période"""
        query = Verification.query
        
        if periode:
            now = datetime.now()
            if periode == "today":
                start_of_day = datetime.combine(now.date(), datetime.min.time())
                query = query.filter(Verification.date_verification >= start_of_day)
            elif periode == "yesterday":
                yesterday = now - timedelta(days=1)
                start_of_day = datetime.combine(yesterday.date(), datetime.min.time())
                end_of_day = datetime.combine(yesterday.date(), datetime.max.time())
                query = query.filter(
                    Verification.date_verification >= start_of_day,
                    Verification.date_verification <= end_of_day
                )
            elif periode == "week":
                week_ago = now - timedelta(days=7)
                query = query.filter(Verification.date_verification >= week_ago)
            elif periode == "month":
                month_ago = now - timedelta(days=30)
                query = query.filter(Verification.date_verification >= month_ago)
            
        
        return query.order_by(
            Verification.date_verification.desc()
        ).limit(limit).all()



    # @staticmethod
    # def get_statistiques_verifications_custom(start_date=None, end_date=None):
    #     """Récupère les statistiques avec filtrage par dates personnalisées"""
    #     query = Verification.query
        
    #     if start_date:
    #         start = datetime.strptime(start_date, '%Y-%m-%d')
    #         query = query.filter(Verification.date_verification >= start)
        
    #     if end_date:
    #         end = datetime.strptime(end_date, '%Y-%m-%d')
    #         # Ajouter 23h59 à la date de fin
    #         end = datetime.combine(end, datetime.max.time())
    #         query = query.filter(Verification.date_verification <= end)
        
    #     total = query.count()
    #     reussies = query.filter(Verification.resultat_donnee == "OK").count()
    #     echouees = query.filter(Verification.resultat_donnee == "ECHEC").count()
    #     externes = query.filter(Verification.resultat_donnee == "EXTERNE").count()

    #     return {
    #         "total": total,
    #         "reussies": reussies,
    #         "echouees": echouees,
    #         "externes": externes,
    #         "start_date": start_date,
    #         "end_date": end_date
    #     }


    @staticmethod
    def get_statistiques_verifications_custom(start_date, end_date):
        """Récupère les statistiques avec filtrage par dates personnalisées"""
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            end = datetime.strptime(end_date, '%Y-%m-%d')
            # Ajouter 23h59 à la date de fin
            end = datetime.combine(end, datetime.max.time())
        except ValueError as e:
            raise ValueError(f"Format de date invalide: {str(e)}")
        
        query = Verification.query.filter(
            Verification.date_verification >= start,
            Verification.date_verification <= end
        )
        
        total = query.count()
        reussies = query.filter(Verification.resultat_donnee == "OK").count()
        echouees = query.filter(Verification.resultat_donnee == "ECHEC").count()
        externes = query.filter(Verification.resultat_donnee == "EXTERNE").count()

        return {
            "total": total,
            "reussies": reussies,
            "echouees": echouees,
            "externes": externes,
            "start_date": start_date,
            "end_date": end_date
        }

    @staticmethod
    def get_stats_verifications_par_lieu_custom(start_date, end_date):
        """Récupère les statistiques par lieu avec filtrage par dates personnalisées"""
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            end = datetime.strptime(end_date, '%Y-%m-%d')
            end = datetime.combine(end, datetime.max.time())
        except ValueError as e:
            raise ValueError(f"Format de date invalide: {str(e)}")
        
        query = Verification.query.filter(
            Verification.date_verification >= start,
            Verification.date_verification <= end
        )
        
        subquery = query.filter(
            Verification.lieu_id.isnot(None)
        ).with_entities(
            Verification.lieu_id,
            func.max(Verification.date_verification).label('derniere_verif')
        ).group_by(
            Verification.lieu_id
        ).subquery()

        results = db.session.query(
            Verification.lieu_id,
            func.count(Verification.id)
        ).join(
            subquery,
            Verification.lieu_id == subquery.c.lieu_id
        ).group_by(
            Verification.lieu_id,
            subquery.c.derniere_verif
        ).order_by(
            subquery.c.derniere_verif.desc()  
        ).limit(10).all()

        data = []
        for lieu_id, total in results:
            lieu = Lieu.query.get(lieu_id)
            if lieu:
                data.append({
                    "lieu": lieu.nom,
                    "total": total
                })

        return data

    @staticmethod
    def get_dernieres_verifications_custom(start_date, end_date, limit=4):
        """Récupère les dernières vérifications avec filtrage par dates personnalisées"""
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            end = datetime.strptime(end_date, '%Y-%m-%d')
            end = datetime.combine(end, datetime.max.time())
        except ValueError as e:
            raise ValueError(f"Format de date invalide: {str(e)}")
        
        return Verification.query.filter(
            Verification.date_verification >= start,
            Verification.date_verification <= end
        ).order_by(
            Verification.date_verification.desc()
        ).limit(limit).all()


# Exportation

    @staticmethod
    def export_verifications_to_excel(verifications):
        """
        Exporte les vérifications vers un fichier Excel avec colonnes conditionnelles
        """
        wb = Workbook()
        
        # Créer 3 feuilles séparées
        ws_ok = wb.active
        ws_ok.title = "Contrôle Autorisé"
        ws_echec = wb.create_sheet("Contrôle Refusé")
        ws_externe = wb.create_sheet("Contrôle Externe")
        
        # Styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Configuration pour chaque type
        sheets_config = {
            'OK': {
                'sheet': ws_ok,
                'headers': ['Date Contrôle', 'Numéro Document', 'Nom Document', 'Prénom Document', 'Utilisateur', 'Lieu'],
                'fill_color': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            },
            'ECHEC': {
                'sheet': ws_echec,
                'headers': ['Date Contrôle', 'Utilisateur', 'Lieu'],
                'fill_color': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            },
            'EXTERNE': {
                'sheet': ws_externe,
                'headers': ['Date Contrôle', 'Nom Externe', 'Prénom Externe', 'Utilisateur', 'Lieu'],
                'fill_color': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            }
        }
        
        # Initialiser les en-têtes pour chaque feuille
        for config in sheets_config.values():
            sheet = config['sheet']
            headers = config['headers']
            
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
        
        # Compteurs de lignes pour chaque feuille
        row_counters = {'OK': 2, 'ECHEC': 2, 'EXTERNE': 2}
        
        # Remplir les données
        for verification in verifications:
            try:
                resultat = verification.resultat_donnee
                
                # Déterminer le type de vérification
                if resultat == 'OK':
                    sheet_type = 'OK'
                elif resultat == 'ECHEC':
                    sheet_type = 'ECHEC'
                else:
                    sheet_type = 'EXTERNE'
                
                sheet = sheets_config[sheet_type]['sheet']
                row = row_counters[sheet_type]
                fill_color = sheets_config[sheet_type]['fill_color']
                
                # Formater la date
                try:
                    date_str = verification.date_verification.strftime('%d/%m/%Y %H:%M:%S')
                except:
                    date_str = str(verification.date_verification)
                
                # Informations utilisateur
                utilisateur_info = "N/A"
                if verification.utilisateur:
                    nom = getattr(verification.utilisateur, 'nom', '') or ""
                    prenom = getattr(verification.utilisateur, 'prenom', '') or ""
                    utilisateur_info = f"{nom} {prenom}".strip() or "N/A"
                
                lieu_info = verification.lieu.nom if verification.lieu else "N/A"
                
                # Remplir selon le type
                if sheet_type == 'OK':
                    numero_doc = "N/A"
                    nom_doc = "N/A"
                    prenom_doc = "N/A"

                    if verification.document:
                        numero_doc = getattr(verification.document, 'numero_document', 'N/A') or "N/A"
                        nom_doc = getattr(verification.document, 'nom', 'N/A') or "N/A"
                        prenom_doc = getattr(verification.document, 'prenom', 'N/A') or "N/A"                    
                        
                    
                    data = [
                        date_str,
                        numero_doc,
                        nom_doc,
                        prenom_doc,
                        utilisateur_info,
                        lieu_info
                    ]
                
                elif sheet_type == 'ECHEC':
                    data = [
                        date_str,
                        utilisateur_info,
                        lieu_info,
                    ]
                
                else:  # EXTERNE
                   
                    nom_ext = "N/A"
                    prenom_ext = "N/A"
                    
                    # Vérifier si ocr_results existe et est une liste
                    if verification.ocr_results:
                        if isinstance(verification.ocr_results, list) and len(verification.ocr_results) > 0:
                            # Prendre le premier résultat OCR
                            ocr = verification.ocr_results[0]
                            nom_ext = getattr(ocr, 'nom_externe', None) or "N/A"
                            prenom_ext = getattr(ocr, 'prenom_externe', None) or "N/A"
                        elif hasattr(verification.ocr_results, 'nom_externe'):
                            # Si c'est un objet unique
                            nom_ext = getattr(verification.ocr_results, 'nom_externe', None) or "N/A"
                            prenom_ext = getattr(verification.ocr_results, 'prenom_externe', None) or "N/A"
                    
                    data = [
                        date_str,
                        nom_ext,
                        prenom_ext,
                        utilisateur_info,
                        lieu_info
                    ]
                
                # Écrire les données dans la feuille
                for col, value in enumerate(data, 1):
                    cell = sheet.cell(row=row, column=col, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    if row % 2 == 0:
                        cell.fill = fill_color
                
                row_counters[sheet_type] += 1
                
            except Exception as e:
                print(f"❌ Erreur lors du traitement de la vérification {verification.id}: {str(e)}")
                import traceback
                traceback.print_exc()
                continue
        
        # Ajuster la largeur des colonnes
        for config in sheets_config.values():
            sheet = config['sheet']
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarder dans un buffer
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer



    @staticmethod
    def get_verifications_for_export(periode=None):
        """
        Récupère les vérifications pour l'export avec leurs relations
        """

        
        query = Verification.query.options(
            joinedload(Verification.utilisateur),
            joinedload(Verification.lieu),
            joinedload(Verification.document),
            joinedload(Verification.ocr_results)
        )
        
        if periode:
            now = datetime.now()
            if periode == "today":
                start_of_day = datetime.combine(now.date(), datetime.min.time())
                query = query.filter(Verification.date_verification >= start_of_day)
            elif periode == "yesterday":
                yesterday = now - timedelta(days=1)
                start_of_day = datetime.combine(yesterday.date(), datetime.min.time())
                end_of_day = datetime.combine(yesterday.date(), datetime.max.time())
                query = query.filter(
                    Verification.date_verification >= start_of_day,
                    Verification.date_verification <= end_of_day
                )
            elif periode == "week":
                week_ago = now - timedelta(days=7)
                query = query.filter(Verification.date_verification >= week_ago)
            elif periode == "month":
                month_ago = now - timedelta(days=30)
                query = query.filter(Verification.date_verification >= month_ago)
        
        return query.order_by(Verification.date_verification.desc()).all()
    
    @staticmethod
    def get_verifications_for_export_custom(start_date, end_date):
        """
        Récupère les vérifications pour l'export avec dates personnalisées
        """
        from models.verification import Verification
        
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            end = datetime.strptime(end_date, '%Y-%m-%d')
            end = datetime.combine(end, datetime.max.time())
        except ValueError as e:
            raise ValueError(f"Format de date invalide: {str(e)}")
        
        return Verification.query.options(
            joinedload(Verification.utilisateur),
            joinedload(Verification.lieu),
            joinedload(Verification.document),
            joinedload(Verification.ocr_results)
        ).filter(
            Verification.date_verification >= start,
            Verification.date_verification <= end
        ).order_by(Verification.date_verification.desc()).all()